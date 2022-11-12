# official Documentation: https://faker.readthedocs.io/en/master/index.html
# Install: pip install Faker

from faker import Faker
from openpyxl import Workbook

fake_data = Faker(['bn_BD', 'hi_IN','en_US']) #'hi_IN','en_US'

# print(fake_data.name())
# print(fake_data.address())
# print(fake_data.email())
wb = Workbook()
ws = wb.active

for i in range(1,200):
    for j in range(1, 8):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()
        ws.cell(row=i, column=4).value = fake_data.phone_number()
        ws.cell(row=i, column=5).value = fake_data.password()
        ws.cell(row=i, column=6).value = fake_data.company()
        ws.cell(row=i, column=7).value = fake_data.credit_card_number()

wb.save("testData.xlsx")
